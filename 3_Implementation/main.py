""" Generates an excel sheet containing all the
required data of a particular Employee using PS Number"""

import openpyxl

from openpyxl.styles import PatternFill, Font, Alignment

from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

from openpyxl.utils import get_column_letter


def rownumber(sheet, ps_number):
    """ Finds the row number of the given PS
    number in the Employee data excel sheet """

    ps_row = 0
    while True:
        for i in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=i, column=2).value
            if cell_value == ps_number:
                ps_row = i
                break
        if ps_row == 0:
            ps_number = int(input("\nPlease Enter Valid PS Number: "))
        else:
            return ps_row


Reading_wb = openpyxl.load_workbook("EmployeeData.xlsx")

PS = int(input("\nEnter the unique PS Number of the"
               " concerned Employee between 88002500 and 88002514: "))

sheet1 = Reading_wb['SemGrades']
sheet2 = Reading_wb['ProgramLang']
sheet3 = Reading_wb['DomainExpertise']
sheet4 = Reading_wb['Hobbies']
sheet5 = Reading_wb['CitiesTravelled']

print(sheet1.max_row)
PS_row = rownumber(sheet1, PS)


# Reads Semester Grades from Excel sheet
# and writes the required data into a new Excel file
New_wb1 = openpyxl.Workbook()

New_wb1['Sheet'].title = "Report"

sh = New_wb1['Report']

sh['D1'] = 'PS Number'
sh['D1'].font = Font(bold=True)
sh['D2'] = PS
sh['E1'] = 'Name'
sh['E1'].font = Font(bold=True)
sh['E2'] = sheet1.cell(PS_row,3).value
sh['F1'] = 'University'
sh['F1'].font = Font(bold=True)
sh['F2'] = sheet1.cell(PS_row,4).value
sh['G1'] = 'Graduation Status'
sh['G1'].font = Font(bold=True)
sh['G2'] = sheet1.cell(PS_row,22).value
if sheet1.cell(PS_row,22).value == 'COMPLETE':
    sh['G2'].fill = PatternFill("solid",fgColor="0000FF00")
else:
    sh['G2'].fill = PatternFill("solid", fgColor="00FF0000")

sh['A4'] = 'Semester'
sh['A4'].fill = PatternFill("solid",fgColor="0033CCCC")
sh['A5'] = 'semester1'
sh['A6'] = 'semester2'
sh['A7'] = 'Year1'
sh['A9'] = 'semester3'
sh['A10'] = 'semester4'
sh['A11'] = 'Year2'
sh['A13'] = 'semester5'
sh['A14'] = 'semester6'
sh['A15'] = 'Year3'
sh['A17'] = 'semester7'
sh['A18'] = 'semester8'
sh['A19'] = 'Year4'

sh['A21'] = 'Final'
sh['A21'].font = Font(bold=True)
for i in range(4,sheet1.max_column+1):
    sh['A'+ str(i)].font = Font(bold=True)

sh['B4'] = 'Grade (CGPA/10)'
sh['B4'].fill = PatternFill("solid",fgColor="0033CCCC")
sh['B4'].font = Font(bold=True)
sh['B5'] = sheet1.cell(PS_row,5).value
sh['B6'] = sheet1.cell(PS_row,6).value
sh['B7'] = sheet1.cell(PS_row,8).value
sh['B9'] = sheet1.cell(PS_row,9).value
sh['B10'] = sheet1.cell(PS_row,10).value
sh['B11'] = sheet1.cell(PS_row,12).value
sh['B13'] = sheet1.cell(PS_row,13).value
sh['B14'] = sheet1.cell(PS_row,14).value
sh['B15'] = sheet1.cell(PS_row,16).value
sh['B17'] = sheet1.cell(PS_row,17).value
sh['B18'] = sheet1.cell(PS_row,18).value
sh['B19'] = sheet1.cell(PS_row,20).value

sh['B21'] = sheet1.cell(PS_row,21).value
sh['B21'].font = Font(bold=True)

for i in range(4,sheet1.max_column+1):
    sh['B'+str(i)].number_format = '0.00'


# Reads Programming Languages known from Excel sheet
# and writes the required data into a new Excel file

sh['C4'] = 'Programming Language'
sh['C4'].fill = PatternFill("solid",fgColor="00CCFFFF")
sh['C5'] = 'C'
sh['C6'] = 'C++'
sh['C7'] = 'Java'
sh['C8'] = 'Python'
sh['C9'] = 'Kotlin'
for i in range(4,sheet2.max_column+1):
    sh['C'+ str(i)].font = Font(bold=True)

sh['D4'] = 'Status'
sh['D4'].fill = PatternFill("solid",fgColor="00CCFFFF")
sh['D4'].font = Font(bold=True)

for i in range(5,sheet2.max_column+1):
    Key = 'D' + str(i)
    sh[Key] = sheet2.cell(PS_row, i).value
    if sh[Key].value == 'YES':
        sh[Key].font = Font(color="00008000")


# Reads Domain Expertise from Excel sheet and
# writes the required data into a new Excel file

sh['E4'] = 'Domain Expertise'
sh['E4'].fill = PatternFill("solid",fgColor="00CC99FF")
sh['E5'] = 'Machine learning'
sh['E6'] = 'Material Science'
sh['E7'] = 'AI'
sh['E8'] = 'Java Script'
sh['E9'] = 'VLSI'
sh['E10'] = 'Semiconductors'
sh['E11'] = 'Networking'
sh['E12'] = 'Embedded'
for i in range(4,sheet3.max_column+1):
    sh['E'+ str(i)].font = Font(bold=True)


sh['F4'] = 'Status'
sh['F4'].font = Font(bold=True)
sh['F4'].fill = PatternFill("solid",fgColor="00CC99FF")

for i in range(5,sheet3.max_column+1):
    Key = 'F' + str(i)
    sh[Key] = sheet3.cell(PS_row, i).value
    if sh[Key].value == 'YES':
        sh[Key].font = Font(color="00008000")


# Reads Hobbies of employees from Excel sheet and
# writes the required data into a new Excel file

sh['G4'] = 'Hobbies'
sh['G4'].fill = PatternFill("solid",fgColor="00FFFFCC")
sh['G5'] = 'Sports'
sh['G6'] = 'Books'
sh['G7'] = 'Singing'
sh['G8'] = 'Dancing'
sh['G9'] = 'Acting'
sh['G10'] = 'Photography'
sh['G11'] = 'Painting'
sh['G12'] = 'Swimming'
for i in range(4,sheet4.max_column+1):
    sh['G'+ str(i)].font = Font(bold=True)


sh['H4'] = 'Status'
sh['H4'].font = Font(bold=True)
sh['H4'].fill = PatternFill("solid",fgColor="00FFFFCC")

for i in range(5,sheet4.max_column+1):
    Key = 'H' + str(i)
    sh[Key] = sheet4.cell(PS_row, i).value
    if sh[Key].value == 'YES':
        sh[Key].font = Font(color="00008000")


# Reads Cities travelled by employees from Excel sheet
# and writes the required data into a new Excel file

sh['I4'] = 'Cities Travelled'
sh['I4'].fill = PatternFill("solid",fgColor="00FF9900")
sh['I5'] = 'Hyderabad'
sh['I6'] = 'Bangalore'
sh['I7'] = 'Chennai'
sh['I8'] = 'Mumbai'
sh['I9'] = 'Delhi'
sh['I10'] = 'Mysore'
sh['I11'] = 'Ahmedabad'
sh['I12'] = 'Vaddodara'
sh['I13'] = 'Baroda'
sh['I14'] = 'Kolkata'
sh['I15'] = 'Coimbatore'
for i in range(4,sheet5.max_column+1):
    sh['I'+ str(i)].font = Font(bold=True)


sh['J4'] = 'Status'
sh['J4'].font = Font(bold=True)
sh['J4'].fill = PatternFill("solid",fgColor="00FF9900")

for i in range(5,sheet5.max_column+1):
    Key = 'j' + str(i)
    sh[Key] = sheet5.cell(PS_row, i).value
    if sh[Key].value == 'YES':
        sh[Key].font = Font(color="00008000")


dim_holder = DimensionHolder(worksheet=sh)
for col in range(sh.min_column, sheet1.max_column + 1):
    dim_holder[get_column_letter(col)] = ColumnDimension(sh, min=col, max=col, width=15)
sh.column_dimensions = dim_holder

rows = range(1, sh.max_row+1)
columns = range(1, sh.max_column+1)
for row in rows:
    for col in columns:
        sh.cell(row, col).alignment = Alignment(horizontal='center',
                                                vertical='center', wrap_text=True)

New_wb1.save("Data.xlsx")
